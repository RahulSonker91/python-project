# package import statement
from SmartApi import SmartConnect #or from SmartApi.smartConnect import SmartConnect
import pyotp
import pandas as pd
from datetime import datetime
from logzero import logger
import openpyxl
import time
import numpy as np

api_key = 'RgOlMNVl'
username = 'MNGZA1072'
pwd = '5315'
smartApi = SmartConnect(api_key)
try:
    token = "AWBIH65LNVEYLFRR7TC53YPHNQ"
    totp = pyotp.TOTP(token).now()
except Exception as e:
    logger.error("Invalid Token: The provided token is not valid.")
    raise e

correlation_id = "abcde"
data = smartApi.generateSession(username, pwd, totp)

if data['status'] == False:
    logger.error(data)
    
else:
    # login api call
    # logger.info(f"You Credentials: {data}")
    authToken = data['data']['jwtToken']
    refreshToken = data['data']['refreshToken']
    # fetch the feedtoken
    feedToken = smartApi.getfeedToken()
    # fetch User Profile
    res = smartApi.getProfile(refreshToken)
    smartApi.generateToken(refreshToken)
    res=res['data']['exchanges']
    
 #Historic api
try:

        wb=openpyxl.load_workbook("NSE_stockTokens.xlsx")
        sh1=wb['list']
        #print(sh1['A1'].value)
        row=sh1.max_row
        column=sh1.max_column
        result=pd.DataFrame(columns=["C2Result","C1Result","P1Result","P2result"])
        for i in range(2,row+1):

                # print(f"fetching {sh1.cell(i,2).value} data for date  {(sh1.cell(i,1).value).date()} ") 
                # print(type(sh1.cell(i,1).value))
                # tradeDate = sh1.cell(i,1).value
                
                # print(datetime.strftime(tradeDate,'%Y-%m-%d'))
                #query to fetch the data from broker  
                historicParam={
                "exchange": "NSE",
                "symboltoken": sh1.cell(i,3).value,
                "interval": "ONE_MINUTE",
                "fromdate": f"{datetime.strftime(sh1.cell(i,1).value,'%Y-%m-%d')} 09:15", 
                "todate": f"{datetime.strftime(sh1.cell(i,1).value,'%Y-%m-%d')} 15:15"
                }

                Pvalue = sh1.cell(i,5).value
                C1Value = sh1.cell(i,8).value
                C2Value = sh1.cell(i,7).value
                P1Value = sh1.cell(i,10).value
                P2Value = sh1.cell(i,11).value

                C1Target = C1Value+Pvalue
                C1Stoploss = C1Value-Pvalue

                C2Target = C2Value+Pvalue
                C2Stoploss = C2Value-Pvalue

                P1Stoploss = P1Value+Pvalue
                P1Target = P1Value-Pvalue
              
                P2Stoploss = P2Value+Pvalue
                P2Target = P2Value-Pvalue
                

                
                # tradeValues = {
                #         'C2' :[C2Value,C2Stoploss,C2Target,C2doTrade],
                #         'C1' :[C1Value,C1Stoploss,C1Target,C1doTrade],
                #         'P1' :[P1Value,P1Stoploss,P1Target,P1doTrade],
                #         'P2' :[P2Value,P2Stoploss,P2Target,P2doTrade]
                # }

                # tradeData = pd.DataFrame(tradeValues,index=['Entry','Stoploss','Target','doTrade'])


                # tradeData.rename(rows={0:'Values'},inplace=True)
                # tradeData.rename(columns={1:'Stoploss'},inplace=True)
                # tradeData.rename(columns={2:'Target'},inplace=True)
                # tradeData.rename(columns={4:'doTrade'},inplace=True)

                tradeData =  pd.DataFrame({
                            'Level': ['C1', 'P1', 'C2', 'P2'],
                            'Entry': [C1Value, P1Value, C2Value, P2Value],
                            'Target': [C1Target, P1Target, C2Target, P2Target],
                            'Stoploss': [C1Stoploss, P1Stoploss, C2Stoploss, P2Stoploss]
                        })
                # print(tradeData)

                # print(C1Value,C2Value,P1Value,P2Value,Pvalue)
        
                #fetchng data from broker
                CS_data = smartApi.getCandleData(historicParam)
                

                # # #Handling fetched data and transfering it to dataframe
                df = pd.DataFrame(CS_data['data'])

                # print(df)

                # # #converting timedata to readable excel format
                df[0] = pd.to_datetime(df[0], format='%Y-%m-%dT%H:%M:%S%z')
                df[0] = df[0].dt.strftime('%Y-%m-%d %H:%M:%S')
                
                #renaming indexes to time,OHLC,values
                df.rename(columns={0: 'datetime'}, inplace=True)
                df.rename(columns={1: 'open'}, inplace=True)
                df.rename(columns={2: 'high'}, inplace=True)
                df.rename(columns={3: 'low'}, inplace=True)
                df.rename(columns={4: 'close'}, inplace=True)
                df.rename(columns={5: 'volume'}, inplace=True)

                df['tradeStatus'] = ""



                # print(datetime.strptime(f"{df['datetime'][0]}",'%Y-%m-%d %H:%M:%S').time() > datetime.strptime("09:10",'%H:%M').time())

                # print("C2Value: " + df.loc[(df['high']>= C1Value) & (df['low']<= C1Value)].head(1)['datetime'])
                # sh1.cell(i,11).value = dt.to_datetime(df.loc[(df['high']>= C2Value) & (df['low']<= C2Value)].head(1)['datetime'])
                # print("P1Value: " + df.loc[(df['high']>= P1Value) & (df['low']<= P1Value)].head(1)['datetime'])
                # print("C1Value: " + df.loc[(df['high']>= P2Value) & (df['low']<= P1Value)].head(1)['datetime'])
               
                # df.loc[(df['high']>= C1Value) & (df['low']<= C1Value),"P1HitTime"] = df['datetime']
                # print(df)
                # sh1.cell(i,9).value = df.loc[(df['high']>= C1Value) & (df['low']<= C1Value)].head(1)['datetime']

                # #writing the data to csv file 
                # with pd.ExcelWriter("NSE_stockTokens.xlsx", engine='openpyxl', mode='a') as writer:
                
                
                #         # Write the DataFrame to a new sheet, using the value from column 1 (A) as the sheet name
                #         df.to_excel(writer, sheet_name=sh1.cell(i, 1).value, index=False)
                # print("C2Value: " + df.loc[(df['high']>= C2Value) & (df['low']<= C2Value)].head(1)['datetime']).strftime('%Y-%m-%d %H:%M:%S')
                # trade=df.map(checkTrade)      
                # print(df)

                def evaluate_trade1(candleStickData,levels):

                                # Sample DataFrames
                                # Candlestick data - replace this with your actual data loading method
                                # candlestick_data = pd.DataFrame({
                                #     'datetime': pd.to_datetime(['2023-10-25 09:15', '2023-10-25 09:35', '2023-10-25 09:40']),
                                #     'open': [100, 102, 104],
                                #     'high': [105, 108, 107],
                                #     'low': [98, 101, 103],
                                #     'close': [102, 106, 104]
                                # })

                        candlestick_data = candleStickData


                                # # Level data - replace this with your actual data loading method
                                # levels = pd.DataFrame({
                                #     'Level': ['C1', 'P1', 'C2', 'P2'],
                                #     'Entry': [104, 99, 106, 101],
                                #     'Target': [110, 95, 112, 97],
                                #     'Stoploss': [98, 103, 100, 105]
                                # })

                        levels =  levels

                        # Initialize trade tracking and results storage
                        trade_results = []
                        level_pairs = {'C1': 'P1', 'P1': 'C1', 'C2': 'P2', 'P2': 'C2'}
                        active_trades = {level: None for level in levels['Level'].unique()}
                        trade_cutoff_time = pd.to_datetime("09:30").time()

                        # Loop through each candlestick
                        for i, candle in candlestick_data.iterrows():
                            candle_time = pd.to_datetime(candle['datetime']).time()
                            
                            # Filter out trades initiated before 09:30
                            if candle_time < trade_cutoff_time:
                                for level, pair in level_pairs.items():
                                    if (levels.set_index('Level').loc[level]['Entry'] >= candle['low'] and 
                                        levels.set_index('Level').loc[level]['Entry'] <= candle['high']):
                                        active_trades[level] = 'Skipped'
                                        active_trades[pair] = 'Skipped'
                                continue

                            # Check for trade initiation at levels
                            for idx, level in levels.iterrows():
                                level_name = level['Level']
                                entry = level['Entry']
                                
                                # Start trade only if it hasn't been initiated or blocked by a pair level
                                if active_trades[level_name] is None and active_trades[level_pairs[level_name]] is None:
                                    if entry >= candle['low'] and entry <= candle['high']:
                                        # Record the trade entry and block counter level
                                        active_trades[level_name] = {
                                            'Entry': candle['datetime'],  # Store as string directly
                                            'Price': entry, 
                                            'Status': 'Open', 
                                            'Target': levels.set_index('Level').loc[level_name]['Target'],
                                            'Stoploss': levels.set_index('Level').loc[level_name]['Stoploss']
                                        }
                                        active_trades[level_pairs[level_name]] = 'Blocked'

                            # Evaluate active trades for either target or stoploss
                            for level_name, trade in active_trades.items():
                                if isinstance(trade, dict) and trade['Status'] == 'Open':
                                    target = trade['Target']
                                    stoploss = trade['Stoploss']
                                    exit_hit = False
                                    
                                    # Evaluate each subsequent candle for exit conditions
                                    for j in range(i, len(candlestick_data)):
                                        exit_candle = candlestick_data.iloc[j]
                                        
                                        # Check for target and stoploss hit based on trade type
                                        if level_name in ['C1', 'C2']:  # Long trades
                                            # Check for target hit
                                            if target <= exit_candle['high']:
                                                # Target hit
                                                trade['Status'] = 'Closed'
                                                trade_results.append({
                                                    'Level': level_name, 
                                                    'Entry Time': trade['Entry'], 
                                                    'Entry Price': trade['Price'],
                                                    'Exit Time': exit_candle['datetime'], 
                                                    'Exit Price': target, 
                                                    'Outcome': 'Target Hit'
                                                })
                                                active_trades[level_name] = 'Closed'  # Close trade after recording result
                                                exit_hit = True
                                                break  # Exit the loop after recording the result
                                            
                                            # Check for stoploss hit
                                            elif stoploss >= exit_candle['low']:
                                                # Stoploss hit
                                                trade['Status'] = 'Closed'
                                                trade_results.append({
                                                    'Level': level_name, 
                                                    'Entry Time': trade['Entry'], 
                                                    'Entry Price': trade['Price'],
                                                    'Exit Time': exit_candle['datetime'], 
                                                    'Exit Price': stoploss, 
                                                    'Outcome': 'Stoploss Hit'
                                                })
                                                active_trades[level_name] = 'Closed'  # Close trade after recording result
                                                exit_hit = True
                                                break  # Exit the loop after recording the result

                                        elif level_name in ['P1', 'P2']:  # Short trades
                                            # Check for target hit
                                            if exit_candle['low'] <= target:
                                                # Target hit
                                                trade['Status'] = 'Closed'
                                                trade_results.append({
                                                    'Level': level_name, 
                                                    'Entry Time': trade['Entry'], 
                                                    'Entry Price': trade['Price'],
                                                    'Exit Time': exit_candle['datetime'], 
                                                    'Exit Price': target, 
                                                    'Outcome': 'Target Hit'
                                                })
                                                active_trades[level_name] = 'Closed'  # Close trade after recording result
                                                exit_hit = True
                                                break  # Exit the loop after recording the result
                                            
                                            # Check for stoploss hit
                                            elif exit_candle['high'] >= stoploss:
                                                # Stoploss hit
                                                trade['Status'] = 'Closed'
                                                trade_results.append({
                                                    'Level': level_name, 
                                                    'Entry Time': trade['Entry'], 
                                                    'Entry Price': trade['Price'],
                                                    'Exit Time': exit_candle['datetime'], 
                                                    'Exit Price': stoploss, 
                                                    'Outcome': 'Stoploss Hit'
                                                })
                                                active_trades[level_name] = 'Closed'  # Close trade after recording result
                                                exit_hit = True
                                                break  # Exit the loop after recording the result

                                    # If the trade is still open after evaluating all candles, check the close condition at 15:15
                                    if not exit_hit and exit_candle['datetime'] >= '2024-10-18 15:15:00':
                                        
                                        close_price = exit_candle['close']
                                        
                                        # Record exit price based on the candle at 15:15
                                        exit_price = close_price
                                        
                                        # Evaluate profit or loss based on trade direction
                                        if level_name in ['C1', 'C2']:  # Long trades
                                            profit_loss = "Profit Target" if exit_price > trade['Price'] else "Loss"
                                        else:  # Short trades
                                            profit_loss = "Profit Target" if exit_price < trade['Price'] else "Loss"

                                        trade['Status'] = 'Closed'
                                        trade_results.append({
                                            'Level': level_name,
                                            'Entry Time': trade['Entry'],
                                            'Entry Price': trade['Price'],
                                            'Exit Time': exit_candle['datetime'],  # Store exit time as string directly
                                            'Exit Price': exit_price,  # Ensure exit price is updated correctly
                                            'Outcome': profit_loss
                                        })
                                        active_trades[level_name] = 'Closed'  # Close trade after recording result

                        # Convert trade results into a DataFrame for analysis
                        trade_results_df = pd.DataFrame(trade_results)

                        if not trade_results_df.empty:
                                print(trade_results_df)
                                # trade_results_df = trade_results_df.append(trade_results_df)

                        # # Initialize trade tracking and results storage
                        # trade_results = []
                        # level_pairs = {'C1': 'P1', 'P1': 'C1', 'C2': 'P2', 'P2': 'C2'}
                        # active_trades = {level: None for level in levels['Level'].unique()}
                        # trade_cutoff_time = pd.to_datetime("09:30").time()

                        # # Loop through each candlestick
                        # for i, candle in candlestick_data.iterrows():
                        #     candle_time = datetime.strptime(f"{candle['datetime']}",'%Y-%m-%d %H:%M:%S').time()
                            
                        #     # Filter out trades initiated before 09:30
                        #     if candle_time < trade_cutoff_time:
                        #         for level, pair in level_pairs.items():
                        #             if (levels.set_index('Level').loc[level]['Entry'] >= candle['low'] and 
                        #                 levels.set_index('Level').loc[level]['Entry'] <= candle['high']):
                        #                 active_trades[level] = 'Skipped'
                        #                 active_trades[pair] = 'Skipped'
                        #         continue

                        #     # Check for trade initiation at levels
                        #     for idx, level in levels.iterrows():
                        #         level_name = level['Level']
                        #         entry = level['Entry']
                                
                        #         # Start trade only if it hasn't been initiated or blocked by a pair level
                        #         if active_trades[level_name] is None and active_trades[level_pairs[level_name]] is None:
                        #             if entry >= candle['low'] and entry <= candle['high']:
                        #                 # Record the trade entry and block counter level
                        #                 active_trades[level_name] = {
                        #                     'Entry': candle['datetime'], 
                        #                     'Price': entry, 
                        #                     'Status': 'Open', 
                        #                     'Target': levels.set_index('Level').loc[level_name]['Target'],
                        #                     'Stoploss': levels.set_index('Level').loc[level_name]['Stoploss']
                        #                 }
                        #                 active_trades[level_pairs[level_name]] = 'Blocked'

                        #     # Evaluate active trades for either target or stoploss
                        #     for level_name, trade in active_trades.items():
                        #         if isinstance(trade, dict) and trade['Status'] == 'Open':
                        #             target = trade['Target']
                        #             stoploss = trade['Stoploss']
                                    
                        #             # Evaluate each subsequent candle for exit conditions
                        #             for j in range(i, len(candlestick_data)):
                        #                 exit_candle = candlestick_data.iloc[j]
                                        
                        #                 # Check for target hit
                        #                 if target <= exit_candle['high']:
                        #                     # Target hit
                        #                     trade['Status'] = 'Closed'
                        #                     trade_results.append({
                        #                         'Level': level_name, 
                        #                         'Entry Time': trade['Entry'], 
                        #                         'Entry Price': trade['Price'],
                        #                         'Exit Time': exit_candle['datetime'], 
                        #                         'Exit Price': target, 
                        #                         'Outcome': 'Target Hit'
                        #                     })
                        #                     active_trades[level_name] = 'Closed'  # Close trade after recording result
                        #                     break  # Exit the loop after recording the result
                                        
                        #                 # Check for stoploss hit
                        #                 elif stoploss >= exit_candle['low']:
                        #                     # Stoploss hit
                        #                     trade['Status'] = 'Closed'
                        #                     trade_results.append({
                        #                         'Level': level_name, 
                        #                         'Entry Time': trade['Entry'], 
                        #                         'Entry Price': trade['Price'],
                        #                         'Exit Time': exit_candle['datetime'], 
                        #                         'Exit Price': stoploss, 
                        #                         'Outcome': 'Stoploss Hit'
                        #                     })
                        #                     active_trades[level_name] = 'Closed'  # Close trade after recording result
                        #                     break  # Exit the loop after recording the result

                        # # Convert trade results into a DataFrame for analysis
                        # trade_results_df = pd.DataFrame(trade_results)

                        # print(trade_results_df)


                                # # Initialize trade tracking and results storage
                                # trade_results = []

                                # # Define trade level pairs and blocking logic
                                # level_pairs = {'C1': 'P1', 'P1': 'C1', 'C2': 'P2', 'P2': 'C2'}
                                # active_trades = {level: None for level in levels['Level'].unique()}

                                # # Convert 09:30 threshold to time for comparison
                                # trade_cutoff_time = pd.to_datetime("09:30").time()

                                # # Loop through each candlestick
                                # for i, candle in candlestick_data.iterrows():
                                #     candle_time = datetime.strptime(f"{candle['datetime']}",'%Y-%m-%d %H:%M:%S').time()
                                    
                                #     # Ignore trades initiated before the 09:30 cutoff time
                                #     if candle_time < trade_cutoff_time:
                                #         for level, pair in level_pairs.items():
                                #             if (levels.set_index('Level').loc[level]['Entry'] >= candle['low'] and 
                                #                 levels.set_index('Level').loc[level]['Entry'] <= candle['high']):
                                #                 # Skip trade for both the level and its counter level
                                #                 active_trades[level] = 'Skipped'
                                #                 active_trades[pair] = 'Skipped'
                                #         continue

                                #     # Check if any level has an entry within the candle's high and low range
                                #     for idx, level in levels.iterrows():
                                #         level_name = level['Level']
                                #         entry = level['Entry']
                                #         target = level['Target']
                                #         stoploss = level['Stoploss']
                                        
                                #         # Check if trade for this level is inactive and not blocked by its pair
                                #         if active_trades[level_name] is None and active_trades[level_pairs[level_name]] is None:
                                #             if entry >= candle['low'] and entry <= candle['high']:
                                #                 # Mark entry and block counter level
                                #                 active_trades[level_name] = {'Entry': candle['datetime'], 'Price': entry, 'Status': 'Open'}
                                #                 active_trades[level_pairs[level_name]] = 'Blocked'

                                #     # Evaluate trade outcomes if there's an active trade
                                #     for level_name, trade in active_trades.items():
                                #         if isinstance(trade, dict) and trade['Status'] == 'Open':
                                #             if target <= candle['high']:
                                #                 # Target hit
                                #                 trade['Status'] = 'Target Hit'
                                #                 trade['Exit Time'] = candle['datetime']
                                #                 trade_results.append({
                                #                     'Level': level_name, 'Entry Time': trade['Entry'], 'Entry Price': trade['Price'],
                                #                     'Exit Time': trade['Exit Time'], 'Exit Price': target, 'Outcome': 'Target Hit'
                                #                 })
                                #             elif stoploss >= candle['low']:
                                #                 # Stoploss hit
                                #                 trade['Status'] = 'Stoploss Hit'
                                #                 trade['Exit Time'] = candle['datetime']
                                #                 trade_results.append({
                                #                     'Level': level_name, 'Entry Time': trade['Entry'], 'Entry Price': trade['Price'],
                                #                     'Exit Time': trade['Exit Time'], 'Exit Price': stoploss, 'Outcome': 'Stoploss Hit'
                                #                 })

                                # # Convert trade results into a DataFrame for analysis
                                # trade_results_df = pd.DataFrame(trade_results)

                                # print(trade_results_df)





                def evaluate_trade(df, entryValue, stop_loss_val, profit_target_val,tradeType):

                        # Calculate stop loss and profit target levels
                        stop_loss_price = stop_loss_val
                        profit_price = profit_target_val
                        tradeEntry = False
                        fromTime = datetime.strptime("09:30",'%H:%M').time()
                        toTime = datetime.strptime("09:30",'%H:%M').time()

                        # print(f"Entry Value-{entryValue}, Stoploss Value:{stop_loss_val},Profit Value:-{profit_target_val}")
                        # Iterate through the rows after entry to evaluate trade outcome
                        for index, row in df.iterrows():
                                # print(row['high'])
                                high = row['high']
                                low = row['low']
                                time = datetime.strptime(f"{row['datetime']}",'%Y-%m-%d %H:%M:%S').time()
                                # print(tradeEntry)
                                # print(f"TradeEntry :{tradeEntry},High Value:{high},Low Value{low},Entry Price:{entryValue},Stoploss Value :{stop_loss_price}, TargetValue :{profit_price}")
                                if low <= entryValue and high >= entryValue and tradeEntry == False and time>=toTime:
                                                tradeEntry = True if not tradeEntry==True else tradeEntry
                                                df.at[index,'tradeStatus'] = "Trade entered"
                                                # print(f'Trade Entered {index},{tradeEntry}')
                                                # return f"Trade entered {row['datetime']}"
                                 
                                 # Check if stop loss is hit
                                if low <= stop_loss_price and high >= stop_loss_price and tradeEntry == True:
                                        tradeEntry=False
                                        df.at[index,'tradeStatus'] = "Stoploss Hit"
                                        # print('Stoploss Hit')
                                        return f"Stop Losshit at {row['datetime']}"

                                # Check if profit target is hit
                                if low <= profit_price and high >= profit_price and tradeEntry == True:
                                        tradeEntry=False
                                        df.at[index,'tradeStatus'] = "Target Hit"
                                        # print('Target Hit')
                                        return f"Profit Targethit at {row['datetime']}"
                                       
                        # return "No result"
                        

                #Example Usage
                
                # C1result = evaluate_trade(df, C1Value, C1Stoploss, C1Target,"C1")
                # # print(C1Value, C1Stoploss, C1Target)
                # # print(f"C1:-{C1result}")
                # # sh1.cell(i,9).value = str(C1result)

                
                # C2result = evaluate_trade(df, C2Value, C2Stoploss, C2Target,"C2")
                # # print(C2Value, C2Stoploss, C2Target)
                # # print(f"C2:-{C2result}")
                # # sh1.cell(i,10).value = str(C2result)

                # P1result = evaluate_trade(df, P1Value, P1Stoploss, P1Target,"P1")
                # # print(f"P1:-{P1result}")
                # # sh1.cell(i,11).value = str(P1result)

                # P2result = evaluate_trade(df, P2Value, P2Stoploss, P2Target,"P2")
                # # print(f"P2:-{P2result}")
                # sh1.cell(i,12).value = str(P2result)
                
                # result.loc[i-2]= [C2result,C1result,P1result,P2result]
                


                ############################################################################################
                evaluate_trade1(df,tradeData)

                

                # print(df.loc[df['tradeStatus']==('Target Hit')])
                # print(df)
 
                ##############################################write candle stick data to excel sheets of particular stock with entry and exit marked##################################
                # with pd.ExcelWriter("NSE_stockTokens.xlsx", engine='openpyxl', mode='a') as writer:
                #         # Write the DataFrame to a new sheet, using the value from column 1 (A) as the sheet name
                #         df.to_excel(writer, sheet_name=sh1.cell(i, 1).value, index=False)

                
                # sh1.cell(i,column+1).value =  df['C1Result'].head(1)
                # sh1.cell(i,column+1).value =  df['C2Result']
                # sh1.cell(i,column+1).value =  df['P1Result'].head(1)
                # sh1.cell(i,column+1).value =  df['P2Result'].head(1)

                # wb.save("NSE_stockTokens_tradeResult.xlsx")
                time.sleep(0.80)

                # def check_trade(row, entry_price, point_value, trade_type):
                #     """Function to check stop loss or profit after entry is made."""
                #     if trade_type == 'long':
                #         profit_price = entry_price + point_value
                #         stop_loss_price = entry_price - point_value
                #         # Check for entry
                #         if row['high'] >= entry_price:
                #             # Check for stop loss or profit
                #             if row['low'] <= stop_loss_price:
                #                 return f"Stop Loss hit at {row['datetime']}, price: {row['low']}"
                #             elif row['high'] >= profit_price:
                #                 return f"Profit Target hit at {row['datetime']}, price: {row['high']}"
                #         return 'Entry not made yet'

                #             elif trade_type == 'short':
                #                 profit_price = entry_price - point_value
                #                 stop_loss_price = entry_price + point_value
                #                 # Check for entry
                #                 if row['low'] <= entry_price:
                #                     # Check for stop loss or profit
                #                     if row['high'] >= stop_loss_price:
                #                         return f"Stop Loss hit at {row['datetime']}, price: {row['high']}"
                #                     elif row['low'] <= profit_price:
                #                         return f"Profit Target hit at {row['datetime']}, price: {row['low']}"
                #                 return 'Entry not made yet'
                #             else:
                #                 raise ValueError("trade_type must be 'long' or 'short'")

                # def evaluate_trade_levels(df, C2, C1, P1, P2, point_value):
                #             # Initialize entry flags
                #             entry_flag_C2 = False
                #             entry_flag_C1 = False
                #             entry_flag_P1 = False
                #             entry_flag_P2 = False
                            
                #             # Define point value for stop loss and profit target
                #             point_value = point_value
                            
                #             results = {'C2_result': [], 'C1_result': [], 'P1_result': [], 'P2_result': []}
                            
                #             for idx, row in df.iterrows():
                #                 # For C2 (long)
                #                 if not entry_flag_C2:
                #                     # Check for entry and further stoploss/profit
                #                     result = check_trade(row, C2, point_value, 'long')
                #                     results['C2_result'].append(result)
                #                     if 'Entry not made yet' not in result:
                #                         entry_flag_C2 = True  # Mark entry for C2
                                
                #                 else:
                #                     # Entry made, now just check for stop loss or profit
                #                     result = check_trade(row, C2, point_value, 'long')
                #                     results['C2_result'].append(result)

                #                 # For C1 (long)
                #                 if not entry_flag_C1:
                #                     result = check_trade(row, C1, point_value, 'long')
                #                     results['C1_result'].append(result)
                #                     if 'Entry not made yet' not in result:
                #                         entry_flag_C1 = True  # Mark entry for C1
                #                 else:
                #                     result = check_trade(row, C1, point_value, 'long')
                #                     results['C1_result'].append(result)

                #                 # For P1 (short)
                #                 if not entry_flag_P1:
                #                     result = check_trade(row, P1, point_value, 'short')
                #                     results['P1_result'].append(result)
                #                     if 'Entry not made yet' not in result:
                #                         entry_flag_P1 = True  # Mark entry for P1
                #                 else:
                #                     result = check_trade(row, P1, point_value, 'short')
                #                     results['P1_result'].append(result)

                #                 # For P2 (short)
                #                 if not entry_flag_P2:
                #                     result = check_trade(row, P2, point_value, 'short')
                #                     results['P2_result'].append(result)
                #                     if 'Entry not made yet' not in result:
                #                         entry_flag_P2 = True  # Mark entry for P2
                #                 else:
                #                     result = check_trade(row, P2, point_value, 'short')
                #                     results['P2_result'].append(result)
    
                #             # Append the results to the dataframe
                #             df['C2_result'] = results['C2_result']
                #             df['C1_result'] = results['C1_result']
                #             df['P1_result'] = results['P1_result']
                #             df['P2_result'] = results['P2_result']
                            
                #             return df


                #             df_with_results = evaluate_trade_levels(df, C2Value, C1Value, P1Value, P2v, Pvalue)

                # print(f"Iteration {i}")
        
        # #printing data
        # print(df)
        for index,row in result.iterrows():
                for col in result.columns:
                        print(f"Value at row {index}, column '{col}': {row[col]}")
        

        # ###################save result to excel file#####################################
        # wb=openpyxl.load_workbook("NSE_stockTokens.xlsx")
        # sh1=wb['list']
        # #print(sh1['A1'].value)
        # row=sh1.max_row
        # column=sh1.max_column
        # print (trade_results_df)
        # for row in result.itertuples(index=True, name="Pandas"):
        #         for i, value in enumerate(row[1:], start=1):  # Skips the index by starting from 1
        #                 sh1.cell(row.Index+2,i+8).value = value
        #                 print(f"Value at row {row.Index}, column '{result.columns[i-1]}': {value}")

        # wb.save("NSE_stockTokens_tradeResult.xlsx")


except Exception as e:
        logger.exception(f"Historic Api failed: {e}")
    #logout
try:
        logout=smartApi.terminateSession('MNGZA1072')
        logger.info("Logout Successfull")
except Exception as e:
        logger.exception(f"Logout failed: {e}")
        print(res)






# def login_broker():
#         try:
#                 api_key = 'RgOlMNVl'
#                 username = 'MNGZA1072'
#                 pwd = '5315'
#                 smartApi = SmartConnect(api_key)
#                 token = "AWBIH65LNVEYLFRR7TC53YPHNQ"
#                 totp = pyotp.TOTP(token).now()
#         except Exception as e:
#                 logger.error("Invalid Token: The provided token is not valid.")
#                 raise e
#                 correlation_id = "abcde"
#                 data = smartApi.generateSession(username, pwd, totp)

#                 if data['status'] == False:
#                         logger.error(data)
#                 else:
#                     # login api call
#                     # logger.info(f"You Credentials: {data}")
#                     authToken = data['data']['jwtToken']
#                     refreshToken = data['data']['refreshToken']
#                     # fetch the feedtoken
#                     feedToken = smartApi.getfeedToken()
#                     # fetch User Profile
#                     res = smartApi.getProfile(refreshToken)
#                     smartApi.generateToken(refreshToken)
#                     res=res['data']['exchanges']

# def logout_broker():
#         try:
#                 logout=smartApi.terminateSession('MNGZA1072')
#                 logger.info("Logout Successfull")
#         except Exception as e:
#                 logger.exception(f"Logout failed: {e}")

#                 correlation_id = "abcde"
#                 data = smartApi.generateSession(username, pwd, totp)

#         if data['status'] == False:
#                 logger.error(data)
    
#         else:
#             # login api call
#             # logger.info(f"You Credentials: {data}")
#             authToken = data['data']['jwtToken']
#             refreshToken = data['data']['refreshToken']
#             # fetch the feedtoken
#             feedToken = smartApi.getfeedToken()
#             # fetch User Profile
#             res = smartApi.getProfile(refreshToken)
#             smartApi.generateToken(refreshToken)
#             res=res['data']['exchanges']
        
# def fetchCandleData(exchange="NSE",symbolToken="7",interval="ONE_MINUTE",fromDate = datetime.today().strftime('%Y-%m-%d'),toDate=datetime.today().strftime('%Y-%m-%d'),startTime = "09:15",EndTime="15:30") :
        
#         #Historic api
#         try:
#                 #query tp fetch the data from broker  
#                 historicParam={
#                 "exchange": exchange,
#                 "symboltoken": symbolToken,
#                 "interval": interval,
#                 "fromdate": fromDate+" "+startTime, 
#                 "todate": toDate+" "+EndTime
#                 }

#                 #fetchng data from broker
#                 CS_data = smartApi.getCandleData(historicParam)
        

#                 #Handling fetched data and transfering it to dataframe
#                 df=pd.DataFrame(CS_data['data'])

#                 #converting timedata to readable excel format
#                 df[0] = pd.to_datetime(df[0], format='%Y-%m-%dT%H:%M:%S%z')
#                 df[0] = df[0].dt.strftime('%Y-%m-%d %H:%M:%S')
                
#                 #renaming indexes to time,OHLC,values
#                 df.rename(columns={0: 'datetime'}, inplace=True)
#                 df.rename(columns={1: 'open'}, inplace=True)
#                 df.rename(columns={2: 'high'}, inplace=True)
#                 df.rename(columns={3: 'low'}, inplace=True)
#                 df.rename(columns={4: 'close'}, inplace=True)
#                 df.rename(columns={5: 'volume'}, inplace=True)

#                 return df
        
#         except Exception as e:
#                 logout_broker()
#                 logger.exception(f"Historic Api failed: {e}")
#                 #logout
        
# wb=openpyxl.load_workbook("NSE_stockTokens.xlsx")
# sh1=wb[wb.active.title]
# row=sh1.max_row
# print(type(sh1))

# login_broker()
# for i in range(1,2):
#         #symbolToken=sh1.cell(i,2)
#         print(f"fetching{sh1.cell(i,2).value}")
#         try:
#                 Sh_data=fetchCandleData()
#                 # Create an ExcelWriter object and append the data to a new sheet
                
#                 with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
#                         writer.book = wb
                
#                 # Write the DataFrame to a new sheet, using the value from column 1 (A) as the sheet name
#                 Sh_data.to_excel(writer, sheet_name=sh1.cell(i, 1).value, index=False)
    
#                 # Save the workbook
#                 writer.save()

#         except Exception as e:
#                 logger.error("Invalid Token: The provided token is not valid.")
#                 raise e
#         finally:
#                 logout_broker()
#                 pass#print(getCandleData())

# logout_broker()        
import pandas as pd
import numpy as np
import requests
import openpyxl


###########################loading a excel sheet to fetch list of stocks############################################
wb=openpyxl.load_workbook("NSE_stockTokens.xlsx")
sh1=wb['list']
#print(sh1['A1'].value)
row=sh1.max_row
#for i in range(1,row+1):
#		print(sh1.cell(i,1).value)

url='https://margincalculator.angelbroking.com/OpenAPI_File/files/OpenAPIScripMaster.json'
d = requests.get(url).json()
token_df = pd.DataFrame.from_dict(d)
token_df['expiry'] = pd.to_datetime(token_df['expiry'],format="mixed").apply(lambda x: x.date())
token_df = token_df.astype({'strike':float})
cols_to_move = ['name','token','symbol']
remaining_cols = [col for col in token_df.columns if col not in cols_to_move]
token_df = token_df[cols_to_move + remaining_cols]
#print(newtoken_df)
#token_df.to_csv('NSE_stock_data.csv')
#print(token_df)
#print(token_df[(token_df['exch_seg']=='NSE')&(token_df['symbol']=='AARTIIND'+'-EQ')]['token'])
#token_df.loc[(token_df[:]['exch_seg']=='NSE' & token_df[:]['name']=='AARTIIND')]


###################################geting stock name to filter tokens from data#########################################################################
stockName=[]
#stockToken=[]
for i in range(2,row+1):
		print(f"{sh1.cell(i,1).value}-{sh1.cell(i,2).value}-{sh1.cell(i,3).value}")
		stockName.append(f"{sh1.cell(i,2).value}-EQ'")
		
		###########################################################save the tokens to excel file###################################################
		#stockToken.append(token_df.loc[(token_df['exch_seg']=='NSE')&(token_df['symbol']==sh1.cell(i,1).value+'-EQ')]['token'].values[0])
		
		sh1.cell(i,3).value = int(float(token_df.loc[(token_df['exch_seg']=='NSE')&(token_df['symbol']==(sh1.cell(i,2).value+'-EQ'))]['token'].values[0]))

		#stockTokens.append((token_df.loc[(token_df['exch_seg']=='NSE')&(token_df['symbol'].isin (stockName))])
		#stockTokens.append(token_df.loc[(token_df['exch_seg'] == 'NSE') & (token_df['symbol'].isin(stockName))])
		#stockName.append(sh1.cell(i,1).value)
		#stock[sh1.cell(i,1).value]=token_df[(token_df['exch_seg']=='NSE')&(token_df['symbol']==stockName+'-EQ')]['token']		
		#print(token_df.loc[(token_df['exch_seg']=='NSE')&(token_df['symbol']==sh1.cell(i,1).value+'-EQ')])
		#print(sh1.cell(i,1).value)
		#print(token_df[(token_df['exch_seg']=='NSE')&(token_df['symbol']==stock+'-EQ')]['token'].astype("string"))
		#sh1.cell(row=i,column=2,value=token_df[(token_df['exch_seg']=='NSE')&(token_df['symbol']==stock+'-EQ')]['token'])
		#print(token_df[(token_df['exch_seg']=='NSE')&(token_df['symbol']==stock+'-EQ')]['token'])
		#print(sh1.cell(i,1).value+" : "+sh1.cell(i,2).value)



wb.save("NSE_stockTokens_fetched_1.xlsx")



#print(stockName)
#wb.save('new_tokens.xlsx')
#token_df.loc[(token_df['exch_seg']=='NSE')&(token_df['symbol'] in stockName)]
#print(token_df.loc[(token_df['exch_seg'] == 'NSE') & (token_df['symbol'].isin(stockName))])
#stockTokens=pd.DataFrame()
#print(type(token_df.loc[(token_df['exch_seg'] == 'NSE') & (token_df['symbol'].isin(stockName))]['token'].tolist()))
#print(stockToken)

#########################################################################################################################
#prepares token list
#tokenLists = token_df.loc[(token_df['exch_seg'] == 'NSE') & (token_df['symbol'].isin(stockName))]['token'].tolist()
#for token in tokenLists:
	#print(token)

####################################saving data to excel######################################
#token_df.loc[(token_df['exch_seg'] == 'NSE') & (token_df['symbol'].isin(stockName))].to_excel("stockTokens.xlsx",sheet_name="stockTokens",index=False)



#print(stockTokens)
#print(token_df)

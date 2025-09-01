import pandas as pd
from datetime import datetime
import openpyxl
import time
import numpy as np

wb=openpyxl.load_workbook("NSE_stockTokens.xlsx")
sh1=wb['list']
#print(sh1['A1'].value)
row=sh1.max_row
column=sh1.max_column
result=pd.DataFrame(columns=["C2Result","C1Result","P1Result","P2result"])
stock_data = pd.read_excel(
        "NSE_stockTokens.xlsx",
        sheet_name="list"
)
print(stock_data.where('symbol' == 'GUJGASLTD'))
# for i in range(2,5):
#     historicParam={
#             "exchange": "NSE",
#             "symboltoken": sh1.cell(i,3).value,
#             "interval": "ONE_MINUTE",
#             "fromdate": f"{datetime.strftime(sh1.cell(i,1).value,'%Y-%m-%d')} 09:15", 
#             "todate": f"{datetime.strftime(sh1.cell(i,1).value,'%Y-%m-%d')} 15:15"
#             }

#     Stock_name = sh1.cell(i,2).value

#     Pvalue = sh1.cell(i,5).value
#     C1Value = sh1.cell(i,8).value
#     C2Value = sh1.cell(i,7).value
#     P1Value = sh1.cell(i,10).value
#     P2Value = sh1.cell(i,11).value

#     C1Target = C1Value+Pvalue
#     C1Stoploss = C1Value-Pvalue

#     C2Target = C2Value+Pvalue
#     C2Stoploss = C2Value-Pvalue

#     P1Stoploss = P1Value+Pvalue
#     P1Target = P1Value-Pvalue
    
#     P2Stoploss = P2Value+Pvalue
#     P2Target = P2Value-Pvalue
    
#     print(f"{Stock_name} : \n C2 - Value:{C2Value} Target:{C2Target } Stoploss:{C2Stoploss}  \n C1 - Value:{C1Value} Target:{C1Target } Stoploss:{C1Stoploss} \n P1- Value:{P1Value} Target:{P1Target} P1Stoploss : {P1Stoploss} \n P2 - Value: {P2Value} Target: {P2Target} - P1Stoploss : {P2Stoploss}")
